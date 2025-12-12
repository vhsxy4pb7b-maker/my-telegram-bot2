"""Excel导出工具"""
import logging
import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import db_operations
from constants import ORDER_STATES

logger = logging.getLogger(__name__)


def create_excel_file(file_path: str, orders: List[Dict], completed_orders: List[Dict] = None, 
                     breach_end_orders: List[Dict] = None, daily_interest: float = 0,
                     daily_summary: Dict = None) -> str:
    """创建Excel文件"""
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # 1. 订单总表工作表
    ws_orders = wb.create_sheet("订单总表", 0)
    
    # 标题
    ws_orders.merge_cells('A1:E1')
    ws_orders['A1'] = "订单总表（有效订单）"
    ws_orders['A1'].font = Font(bold=True, size=14)
    ws_orders['A1'].alignment = center_align
    
    # 表头
    headers = ['时间', '订单号', '金额', '状态', '利息记录']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_orders.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # 数据行
    row_idx = 3
    for order in orders:
        date_str = order.get('date', '')[:10] if order.get('date') else '未知'
        order_id = order.get('order_id', '未知')
        amount = order.get('amount', 0)
        state = ORDER_STATES.get(order.get('state', ''), order.get('state', '未知'))
        
        # 订单基本信息行
        ws_orders.cell(row=row_idx, column=1, value=date_str).border = border
        ws_orders.cell(row=row_idx, column=2, value=order_id).border = border
        ws_orders.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
        ws_orders.cell(row=row_idx, column=3).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=3).alignment = right_align
        ws_orders.cell(row=row_idx, column=4, value=state).border = border
        
        # 获取利息记录（从传入的订单数据中获取，如果订单有interests字段）
        interests = order.get('interests', [])
        
        if interests:
            interest_text = "\n".join([
                f"{interest.get('date', '')[:10] if interest.get('date') else '未知'}: "
                f"{float(interest.get('amount', 0)):,.2f}"
                for interest in interests
            ])
            ws_orders.cell(row=row_idx, column=5, value=interest_text).border = border
        else:
            ws_orders.cell(row=row_idx, column=5, value="无").border = border
        
        row_idx += 1
    
    # 汇总行
    if daily_interest > 0:
        ws_orders.merge_cells(f'A{row_idx}:D{row_idx}')
        ws_orders.cell(row=row_idx, column=1, value="当日利息汇总:").font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=5, value=float(daily_interest)).number_format = '#,##0.00'
        ws_orders.cell(row=row_idx, column=5).font = Font(bold=True)
        ws_orders.cell(row=row_idx, column=5).alignment = right_align
    
    # 调整列宽
    ws_orders.column_dimensions['A'].width = 12
    ws_orders.column_dimensions['B'].width = 15
    ws_orders.column_dimensions['C'].width = 15
    ws_orders.column_dimensions['D'].width = 10
    ws_orders.column_dimensions['E'].width = 30
    
    # 2. 已完成订单工作表
    if completed_orders:
        ws_completed = wb.create_sheet("已完成订单")
        ws_completed.merge_cells('A1:D1')
        ws_completed['A1'] = "已完成订单（当日）"
        ws_completed['A1'].font = Font(bold=True, size=14)
        ws_completed['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '完成时间']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_completed.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in completed_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            updated_at = order.get('updated_at', '')[:19] if order.get('updated_at') else '未知'
            
            ws_completed.cell(row=row_idx, column=1, value=date_str).border = border
            ws_completed.cell(row=row_idx, column=2, value=order_id).border = border
            ws_completed.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_completed.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_completed.cell(row=row_idx, column=3).alignment = right_align
            ws_completed.cell(row=row_idx, column=4, value=updated_at).border = border
            
            row_idx += 1
        
        ws_completed.column_dimensions['A'].width = 12
        ws_completed.column_dimensions['B'].width = 15
        ws_completed.column_dimensions['C'].width = 15
        ws_completed.column_dimensions['D'].width = 20
    
    # 3. 违约完成订单工作表
    if breach_end_orders:
        ws_breach = wb.create_sheet("违约完成订单")
        ws_breach.merge_cells('A1:D1')
        ws_breach['A1'] = "违约完成订单（当日有变动）"
        ws_breach['A1'].font = Font(bold=True, size=14)
        ws_breach['A1'].alignment = center_align
        
        headers = ['时间', '订单号', '金额', '完成时间']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_breach.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 3
        for order in breach_end_orders:
            date_str = order.get('date', '')[:10] if order.get('date') else '未知'
            order_id = order.get('order_id', '未知')
            amount = order.get('amount', 0)
            updated_at = order.get('updated_at', '')[:19] if order.get('updated_at') else '未知'
            
            ws_breach.cell(row=row_idx, column=1, value=date_str).border = border
            ws_breach.cell(row=row_idx, column=2, value=order_id).border = border
            ws_breach.cell(row=row_idx, column=3, value=float(amount) if amount else 0).border = border
            ws_breach.cell(row=row_idx, column=3).number_format = '#,##0.00'
            ws_breach.cell(row=row_idx, column=3).alignment = right_align
            ws_breach.cell(row=row_idx, column=4, value=updated_at).border = border
            
            row_idx += 1
        
        ws_breach.column_dimensions['A'].width = 12
        ws_breach.column_dimensions['B'].width = 15
        ws_breach.column_dimensions['C'].width = 15
        ws_breach.column_dimensions['D'].width = 20
    
    # 4. 日切数据汇总工作表
    if daily_summary:
        ws_summary = wb.create_sheet("日切数据汇总")
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = "日切数据汇总"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A1'].alignment = center_align
        
        summary_data = [
            ['新增订单数', daily_summary.get('new_orders_count', 0)],
            ['新增订单金额', daily_summary.get('new_orders_amount', 0.0)],
            ['完结订单数', daily_summary.get('completed_orders_count', 0)],
            ['完结订单金额', daily_summary.get('completed_orders_amount', 0.0)],
            ['违约完成数', daily_summary.get('breach_end_orders_count', 0)],
            ['违约完成金额', daily_summary.get('breach_end_orders_amount', 0.0)],
            ['当日利息', daily_summary.get('daily_interest', 0.0)],
            ['公司开销', daily_summary.get('company_expenses', 0.0)],
            ['其他开销', daily_summary.get('other_expenses', 0.0)],
            ['总开销', daily_summary.get('company_expenses', 0.0) + daily_summary.get('other_expenses', 0.0)],
        ]
        
        row_idx = 3
        for label, value in summary_data:
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=1).border = border
            if isinstance(value, float):
                ws_summary.cell(row=row_idx, column=2, value=value).number_format = '#,##0.00'
                ws_summary.cell(row=row_idx, column=2).alignment = right_align
            else:
                ws_summary.cell(row=row_idx, column=2, value=value).alignment = center_align
            ws_summary.cell(row=row_idx, column=2).border = border
            row_idx += 1
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 20
    
    # 保存文件
    wb.save(file_path)
    return file_path


async def export_orders_to_excel(orders: List[Dict], completed_orders: List[Dict] = None,
                                breach_end_orders: List[Dict] = None, daily_interest: float = 0,
                                daily_summary: Dict = None) -> str:
    """导出订单到Excel文件（异步版本）"""
    import asyncio
    import tempfile
    
    # 为每个订单获取利息记录
    orders_with_interests = []
    for order in orders:
        order_id = order.get('order_id')
        if order_id:
            try:
                interests = await db_operations.get_all_interest_by_order_id(order_id)
                order_copy = order.copy()
                order_copy['interests'] = interests
                orders_with_interests.append(order_copy)
            except Exception as e:
                logger.error(f"获取订单 {order_id} 的利息记录失败: {e}")
                order_copy = order.copy()
                order_copy['interests'] = []
                orders_with_interests.append(order_copy)
        else:
            orders_with_interests.append(order)
    
    # 创建临时文件
    temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"订单报表_{timestamp}.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    
    # 在事件循环中运行同步函数
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        create_excel_file,
        file_path, orders_with_interests, completed_orders, breach_end_orders, daily_interest, daily_summary
    )
    
    return file_path

